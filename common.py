#!/usr/bin/env python3
"""
    Common Functions to use with Intersight API
    Libraries: pip install requests jsonpath-ng openpyxl flatten-json
"""
import sys
import os
import json
import re
import time
import requests
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def get_token(client_id, client_secret):
    """ Get oAuth Token """
    token_url="https://intersight.com/iam/token"
    client_auth = requests.auth.HTTPBasicAuth(client_id, client_secret)
    post_data = {"grant_type": "client_credentials"}
    response = requests.post(url=token_url,
                            auth=client_auth,
                            data=post_data)
    if response.status_code != 200:
        print("Failed to obtain token from the OAuth 2.0 server", file=sys.stderr)
        sys.exit(1)
    print("Successfuly obtained a new token")
    json_data = response.json()
    token = json_data["access_token"]
    return token


def get_api_data(client_id, client_secret, token, api_url):
    """ Get API Endpoint Data """
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url=api_url, headers=headers)
    data = response.json()
    if	response.status_code == 401:
        print("-> Existing Token Expired. Generating a new one!")
        token = get_token(client_id, client_secret)
        get_api_data(client_id, client_secret, token, api_url)
    if response.status_code == 429:
        wait = response.headers.get("Retry-After", 600)
        print(f"-> got {response.status_code} from {api_url}. retrying after {wait}s")
        time.sleep(int(wait))
    elif response.status_code == requests.codes.ok:
        return data
    else:
        print(f"-> got {response.status_code} from {api_url}")


def get_count(client_id, client_secret, token, api_url):
    """
        Return count of API Endpoint objects
    """
    response = get_api_data(client_id, client_secret, token, api_url)
    total_count = response["Count"]
    return total_count


def get_all(client_id, client_secret, token, api_url, total_count):
    """
        Pagination code to get all the objects
    """
    increment = 1000
    skip = 0
    data = []
    while (skip <= total_count):
        if "?" in api_url:
            api_path = f"{api_url}&$top=1000&$skip={skip}"
        else:
            api_path = f"{api_url}?$top=1000&$skip={skip}"
        response = get_api_data(client_id, client_secret, token, api_path)
        data.extend(response["Results"])
        skip += increment
    return data


def get_data(client_id, client_secret, token, api_count_url, api_url):
    """
        Get Object Total Count and finally all the data
    """
    # Get Endpoint Data Count
    total_count = get_count(client_id, client_secret, token, api_count_url)
    # Get All Data
    data = get_all(client_id, client_secret, token, api_url, total_count)
    return data


def get_excel(file_name, sheet_name):
    """
        Verify if inventory file exists. If not, create one.
        Add Sheet Name based on provided sheet_name
        Return active sheet
    """
    if os.path.exists(file_name) and os.path.isfile(file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        if len(workbook.sheetnames) == 1 and sheet.title == "Sheet":
            sheet.title = sheet_name
            workbook.save(file_name)
        elif sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            workbook.save(file_name)
    else:
        workbook = Workbook()
        workbook.save(file_name)
        sheet = workbook.active
        sheet.title = sheet_name
        workbook.save(file_name)
    return workbook, sheet
       

def add_header_row(sheet, header_list):
    """
        Add Headers Row to the Sheet
    """
    for column, value in enumerate(header_list, start=1):
        sheet.cell(row=1, column=column, value=value)
        cell = sheet.cell(row=1, column=column)
        custom_font = Font(name='Calibri', size=18)  
        blue_fill = PatternFill(start_color='66ccff', end_color='66ccff', fill_type='solid')
        cell.font = custom_font
        cell.fill = blue_fill


# def add_cell_data(sheet, data):
#     """
#         Add Data in Sheet Cells
#     """
#     for row, item in enumerate(data, start=2):
#         for column in range(1, len(list(item.keys()))+1):
#             sheet.cell(row=row,column=column, value=list(item.values())[column-1])
#             cell = sheet.cell(row=row, column=column)
#             custom_font = Font(name='Calibri', size=14)  
#             cell.font = custom_font
#             # cell.alignment=Alignment(vertical='center')

def add_cell_data(sheet, header_list, data):
    """
        Add Data in Sheet Cells
    """
    for row, item in enumerate(data, start=2):
        for column_name in item.keys():
            index = header_list.index(column_name)
            sheet.cell(row=row,column=index+1, value=item[column_name])
            cell = sheet.cell(row=row, column=index+1)
            custom_font = Font(name='Calibri', size=14)  
            cell.font = custom_font



def write_to_excel(file_name, sheet_name, header_list, data):
    """
        Add Data to an Excel file under the provided sheet name
    """
    # Get or Create Workbook and Sheet Name
    workbook, sheet = get_excel(file_name, sheet_name)

    # Write Sheet Headers
    # column_len = len(header_list)
    add_header_row(sheet, header_list)

    # Write Sheet Cells
    # add_cell_data(sheet, column_len, data)
    add_cell_data(sheet, header_list, data)
    
    # Save and Close Workbook
    workbook.save(file_name)
    workbook.close()


def flatten_json(y):
    """
        Flatten Json data
    """
    out = {}

    def flatten(x, name=''):
        if type(x) is dict:
            for a in x:
                flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


def parse_data(data):
    """
        Apply flatten_json function on the data
    """
    parsed_data = []
    for item in data:
        flat_item = flatten_json(item)
        parsed_data.append(flat_item)
    return parsed_data


def remove_parameters(parsed_data):
    """
        Remove Default parameters from output
    """
    data = parsed_data
    parameters = ["ObjectType", "ClassId", "Parent_ObjectType", "Parent_ClassId", "Board_ObjectType", "Board_ClassId", "NetworkElement_ClassId", "NetworkElement_ObjectType", "RegisteredDevice_Moid", "RegisteredDevice_ClassId", "RegisteredDevice_ObjectType", "Contract_ClassId", "Contract_ObjectType", "Contract_BillTo_ClassId", "Contract_BillTo_ObjectType", "Contract_BillTo_Address1", "Contract_BillTo_Address2", "Contract_BillTo_Address3", "Contract_BillTo_City", "Contract_BillTo_Country", "Contract_BillTo_County", "Contract_BillTo_Location", "Contract_BillTo_Name", "Contract_BillTo_PostalCode", "Contract_BillTo_Province", "Contract_BillTo_State", "Contract_BillToGlobalUltimate_ClassId", "Contract_BillToGlobalUltimate_ObjectType", "Contract_BillToGlobalUltimate_Id", "Contract_BillToGlobalUltimate_Name", "Source_ClassId", "Source_ObjectType", "Source_Moid", "Source_Name", "Source_PlatformType"]
    for item in data:
        for p in parameters:
            if p in item:
                item.pop(p)
    return data


def auto_size_columns(file_name, sheet_name):
    """
        Iterate over all columns and adjust their widths
    """
    workbook, sheet = get_excel(file_name, sheet_name)
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column_letter].width = adjusted_width
    # Save and Close Workbook
    workbook.save(file_name)
    workbook.close()


def find_empty_slots(data):
    """
        Find Empty Slots in Chassis
        Output Domain Id, Chassis Id, Slot Id.
    """
    occupied_slots = {}
    for blade in data:
        server_name = blade["Name"]
        parsed_name = re.search(r"^(.+)-(\d+)-(\d)", server_name)
        domain_id = parsed_name.group(1)
        chassis_id = parsed_name.group(2)
        server_id = parsed_name.group(3)
        if domain_id not in occupied_slots.keys():
            occupied_slots[domain_id] = {}
        if chassis_id not in occupied_slots[domain_id].keys():
            occupied_slots[domain_id][chassis_id] = []
        if server_id not in occupied_slots[domain_id][chassis_id]:
            occupied_slots[domain_id][chassis_id].append(server_id)
    
    empty_slots = {}
    for domain_id,chassis in occupied_slots.items():
        for chassis_id,servers in chassis.items():
            for i in range(1,9):
                if str(i) not in servers:
                    domain_id = domain_id
                    chassis_id = chassis_id
                    server_id = str(i)
                    if domain_id not in empty_slots.keys():
                        empty_slots[domain_id] = {}
                    if chassis_id not in empty_slots[domain_id].keys():
                        empty_slots[domain_id][chassis_id] = []
                    if server_id not in empty_slots[domain_id][chassis_id]:
                        empty_slots[domain_id][chassis_id].append(server_id)
    
    parsed_data = []
    for domain_id,chassis in empty_slots.items():
        for chassis_id,servers in chassis.items():
            server_ids = ', '.join(servers)
            data_dict = {
                "domain_id": domain_id,
                "chassis_id": chassis_id,
                "server_slots": server_ids
            }
            parsed_data.append(data_dict)
    return parsed_data


def create_hyperlinks_sheet(file_name):
    """
        Create a Hyperlinks sheet pointing to all the Sheets in excel
    """
    sheet_name = "Hyperlinks"
    workbook, sheet = get_excel(file_name, sheet_name)

    # Set Header Row
    sheet.cell(row=1, column=1, value="Hyperlinks")
    cell = sheet.cell(row=1, column=1)
    custom_font = Font(name='Calibri', size=18)  
    blue_fill = PatternFill(start_color='66ccff', end_color='66ccff', fill_type='solid')
    cell.font = custom_font
    cell.fill = blue_fill

    # Add Hyperlinks
    sheets = workbook.sheetnames
    if "Hyperlinks" in sheets:
        sheets.remove('Hyperlinks')

    for i,sheet_name in enumerate(sheets, start=2):
        cell_font = Font(name="Calibri", underline="single", size=18, color="0066cc")
        cell_value = sheet_name.upper()
        link_value = f"#{sheet_name}!A{i}"
        sheet.cell(row=i, column=1, value=cell_value).hyperlink = link_value
        sheet.cell(row=i, column=1).font = cell_font

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save and close Workbook
    workbook.save(file_name)
    workbook.close()

def set_default_sheet(file_name, sheet_name):
    workbook = load_workbook(filename=file_name)

    # Get the sheet's current position (index)
    sheet_index = workbook.sheetnames.index(sheet_name)
    new_position = - sheet_index
    
    # Move sheet to Index 0
    workbook.move_sheet(sheet_name, new_position)

    # Set Active sheet
    workbook.active = workbook[sheet_name]

    # Save the changes
    workbook.save(file_name) 
    workbook.close()

def get_licenses(data):
    """
        Get Server Licenses
    """
    license_data = []
    for d in data:
        license_dict = {}
        license_dict['Name'] = d['Name']
        license_dict['Serial'] = d['Serial']
        license_dict['Model'] = d['Model']
        for k,v in d.items():
            if "Intersight.LicenseTier" == v:
                prefix = k.split("_Key")[0]
                license_key = f"{prefix}_Value"
                license_dict["LicenseTier"] = d[license_key]
        license_data.append(license_dict)
    return license_data


def get_sp_policies(data):
    """
        Get Server Profile and its associated Policies
    """
    sp_data = []
    for d in data:
        sp_dict = {}
        sp_dict['SP_Name'] = d['Name']
        sp_dict['SP_Moid'] = d['Moid']
        sp_dict['TargetPlatform'] = d['TargetPlatform']
        if 'AssociatedServer' in d.keys():
            if d['AssociatedServer'] == None:
                sp_dict['Server_Name'] = ""
                sp_dict['Server_Serial'] = ""
                sp_dict['Server_Model'] = ""
        else:
            if 'AssociatedServer_Name' in d.keys():
                sp_dict['Server_Name'] = d['AssociatedServer_Name']
            # if 'AssociatedServer_Serial' in d.keys():
                sp_dict['Server_Serial'] = d['AssociatedServer_Serial']
            # if 'AssociatedServer_Model' in d.keys():
                sp_dict['Server_Model'] = d['AssociatedServer_Model']
        for k,v in d.items():
            if "ClassId" in k and "PolicyBucket" in k:
                prefix = k.split("_ClassId")[0]
                policy_key = f"{prefix}_Name"
                sp_dict[v] = d[policy_key]                                                         
        sp_data.append(sp_dict)
    return sp_data

def get_vnic_ethifs(client_id, client_secret, token, fi_veth_data):
    """
        Get Virtual Ethernet Interfaces
    """
    api_count_url = "https://intersight.com/api/v1/vnic/EthIfs?$filter=LcpVnic ne 'null'&$count=True"
    api_url = "https://intersight.com/api/v1/vnic/EthIfs?$filter=LcpVnic ne 'null'&$expand=Profile($select=Name,AssociatedServer%3B$expand=AssociatedServer($select=Name,Model,Serial)),EthQosPolicy($select=Mtu,Cos,Priority),FabricEthNetworkGroupPolicy($select=VlanSettings),LcpVnic($select=LanConnectivityPolicy%3B$expand=LanConnectivityPolicy($select=Name))&$select=Name,MacAddress,FailoverEnabled,VifId,StandbyVifId,Placement,Profile,EthQosPolicy,FabricEthNetworkGroupPolicy,LcpVnic"

    # Intersight API Nested Data        
    data = get_data(client_id, client_secret, token, api_count_url, api_url)
    
    if data:
        # Flattened Data
        semi_parsed_data = parse_data(data)
        # Remove Parameters from Parsed Data before Writing
        parsed_data = remove_parameters(semi_parsed_data)

    vnic_data = []
    for vnic in parsed_data:
        vnic_dict = {}
        vnic_dict["Cos"]                   = vnic["EthQosPolicy_Cos"]
        vnic_dict["Mtu"]                   = vnic["EthQosPolicy_Mtu"]
        vnic_dict["Priority"]              = vnic["EthQosPolicy_Priority"]
        vnic_dict["AllowedVlans"]          = vnic["FabricEthNetworkGroupPolicy_0_VlanSettings_AllowedVlans"]
        vnic_dict["NativeVlan"]            = vnic["FabricEthNetworkGroupPolicy_0_VlanSettings_NativeVlan"]
        vnic_dict["QinqEnabled"]           = vnic["FabricEthNetworkGroupPolicy_0_VlanSettings_QinqEnabled"]
        vnic_dict["QinqVlan"]              = vnic["FabricEthNetworkGroupPolicy_0_VlanSettings_QinqVlan"]
        vnic_dict["FailoverEnabled"]       = vnic["FailoverEnabled"]
        vnic_dict["LanConnectivityPolicy"] = vnic["LcpVnic_LanConnectivityPolicy_Name"]
        vnic_dict["MacAddress"]            = vnic["MacAddress"]
        vnic_dict["Moid"]                  = vnic["Moid"]
        vnic_dict["Name"]                  = vnic["Name"]
        vnic_dict["AutoPciLink"]           = vnic["Placement_AutoPciLink"]
        vnic_dict["AutoSlotId"]            = vnic["Placement_AutoSlotId"]
        vnic_dict["Slot_Id"]               = vnic["Placement_Id"]
        vnic_dict["PciLink"]               = vnic["Placement_PciLink"]
        vnic_dict["PciLinkAssignmentMode"] = vnic["Placement_PciLinkAssignmentMode"]
        vnic_dict["SwitchId"]              = vnic["Placement_SwitchId"]
        vnic_dict["Uplink"]                = vnic["Placement_Uplink"]
        vnic_dict["Profile_Name"]          = vnic["Profile_Name"]
        vnic_dict["StandbyVifId"]          = vnic["StandbyVifId"]
        vnic_dict["VifId"]                 = vnic["VifId"]
        for veth in fi_veth_data:
            server_serial = veth["Description"].split(":")[1]
            vnic_name = (veth["Description"].split()[3]).split(",")[0]
            if "Profile_AssociatedServer" not in vnic.keys():
                if (vnic["Profile_AssociatedServer_Serial"] == server_serial) and (vnic["Name"] == vnic_name) and (veth["VethId"] == vnic["VifId"]):
                    vnic_dict["AssociatedServer_Model"]  = vnic["Profile_AssociatedServer_Model"]
                    vnic_dict["AssociatedServer_Name"]   = vnic["Profile_AssociatedServer_Name"]
                    vnic_dict["AssociatedServer_Serial"] = vnic["Profile_AssociatedServer_Serial"]
                    vnic_dict["BoundInterfaceDn"]        = veth["BoundInterfaceDn"]
                    vnic_dict["Veth_Description"]        = veth["Description"]
                    vnic_dict["FI_AdminEvacState"]       = veth["NetworkElement_AdminEvacState"]
                    vnic_dict["FI_ManagementMode"]       = veth["NetworkElement_ManagementMode"]
                    vnic_dict["FI_Model"]                = veth["NetworkElement_Model"]
                    vnic_dict["FI_OperEvacState"]        = veth["NetworkElement_OperEvacState"]
                    vnic_dict["FI_Operability"]          = veth["NetworkElement_Operability"]
                    vnic_dict["FI_Serial"]               = veth["NetworkElement_Serial"]
                    vnic_dict["FI_SwitchId"]             = veth["NetworkElement_SwitchId"]
                    vnic_dict["FI_SwitchProfileName"]    = veth["NetworkElement_SwitchProfileName"]
                    vnic_dict["OperReason"]              = veth["OperReason"]
                    vnic_dict["OperState"]               = veth["OperState"]
                    vnic_dict["PinnedInterfaceDn"]       = veth["PinnedInterfaceDn"]
                    vnic_dict["VethId"]                  = veth["VethId"]
            else:
                if (vnic["Name"] == vnic_name) and (veth["VethId"] == vnic["VifId"]):
                    vnic_dict["AssociatedServer_Model"]  = ""
                    vnic_dict["AssociatedServer_Name"]   = ""
                    vnic_dict["AssociatedServer_Serial"] = ""
                    vnic_dict["BoundInterfaceDn"]        = veth["BoundInterfaceDn"]
                    vnic_dict["Veth_Description"]        = veth["Description"]
                    vnic_dict["FI_AdminEvacState"]       = veth["NetworkElement_AdminEvacState"]
                    vnic_dict["FI_ManagementMode"]       = veth["NetworkElement_ManagementMode"]
                    vnic_dict["FI_Model"]                = veth["NetworkElement_Model"]
                    vnic_dict["FI_OperEvacState"]        = veth["NetworkElement_OperEvacState"]
                    vnic_dict["FI_Operability"]          = veth["NetworkElement_Operability"]
                    vnic_dict["FI_Serial"]               = veth["NetworkElement_Serial"]
                    vnic_dict["FI_SwitchId"]             = veth["NetworkElement_SwitchId"]
                    vnic_dict["FI_SwitchProfileName"]    = veth["NetworkElement_SwitchProfileName"]
                    vnic_dict["OperReason"]              = veth["OperReason"]
                    vnic_dict["OperState"]               = veth["OperState"]
                    vnic_dict["PinnedInterfaceDn"]       = veth["PinnedInterfaceDn"]
                    vnic_dict["VethId"]                  = veth["VethId"]
        vnic_data.append(vnic_dict)
    return vnic_data

def get_vhba_fcifs(client_id, client_secret, token, fi_vfc_data):
    """
        Get vHBA Interfaces
    """
    api_count_url = "https://intersight.com/api/v1/vnic/FcIfs?$filter=ScpVhba ne 'null'&$count=True"
    api_url = "https://intersight.com/api/v1/vnic/FcIfs?$filter=ScpVhba ne 'null'&$expand=FcAdapterPolicy($select=IoThrottleCount,LunCount,LunQueueDepth),FcNetworkPolicy($select=Name,VsanSettings),WwpnPool($select=Name),FcQosPolicy($select=Burst,Cos,Name,Priority,RateLimit),Profile($select=Name,AssociatedServer%3B$expand=AssociatedServer($select=Name,Model,Serial)),ScpVhba($select=SanConnectivityPolicy%3B$expand=SanConnectivityPolicy($select=Name))&$select=Name,Order,Placement,FcAdapterPolicy,FcNetworkPolicy,FcQosPolicy,Profile,ScpVhba,Type,VifId,Wwpn,WwpnAddressType,WwpnPool"

    # Intersight API Nested Data        
    data = get_data(client_id, client_secret, token, api_count_url, api_url)
    
    if data:
        # Flattened Data
        semi_parsed_data = parse_data(data)
        # Remove Parameters from Parsed Data before Writing
        parsed_data = remove_parameters(semi_parsed_data)

        vhba_data = []
        for vhba in parsed_data:
            vhba_dict = {}
            vhba_dict["IoThrottleCount"]                = vhba["FcAdapterPolicy_IoThrottleCount"]
            vhba_dict["LunCount"]                       = vhba["FcAdapterPolicy_LunCount"]
            vhba_dict["LunQueueDepth"]                  = vhba["FcAdapterPolicy_LunQueueDepth"]
            vhba_dict["FcNetPolicy_Name"]               = vhba["FcNetworkPolicy_Name"]
            vhba_dict["FcNetPolicy_Vsan_DefaultVlanId"] = vhba["FcNetworkPolicy_VsanSettings_DefaultVlanId"]
            vhba_dict["FcNetPolicy_Vsan_Id"]            = vhba["FcNetworkPolicy_VsanSettings_Id"]
            vhba_dict["Qos_Burst"]                      = vhba["FcQosPolicy_Burst"]
            vhba_dict["Qos_Cos"]                        = vhba["FcQosPolicy_Cos"]
            vhba_dict["QosPolicy_Name"]                 = vhba["FcQosPolicy_Name"]
            vhba_dict["Qos_Priority"]                   = vhba["FcQosPolicy_Priority"]
            vhba_dict["Qos_RateLimit"]                  = vhba["FcQosPolicy_RateLimit"]
            vhba_dict["Moid"]                           = vhba["Moid"]
            vhba_dict["Name"]                           = vhba["Name"]
            vhba_dict["AutoPciLink"]                    = vhba["Placement_AutoPciLink"]
            vhba_dict["AutoSlotId"]                     = vhba["Placement_AutoSlotId"]
            vhba_dict["Placement_Id"]                   = vhba["Placement_Id"]
            vhba_dict["PciLink"]                        = vhba["Placement_PciLink"]
            vhba_dict["PciLinkAssignmentMode"]          = vhba["Placement_PciLinkAssignmentMode"]
            vhba_dict["SwitchId"]                       = vhba["Placement_SwitchId"]
            vhba_dict["Uplink"]                         = vhba["Placement_Uplink"]
            vhba_dict["Profile_Name"]                   = vhba["Profile_Name"]
            vhba_dict["SanConnPolicy_Name"]             = vhba["ScpVhba_SanConnectivityPolicy_Name"]
            vhba_dict["VifId"]                          = vhba["VifId"]
            vhba_dict["WwpnPool_Name"]                  = vhba["WwpnPool_Name"]
            if "Profile_AssociatedServer" not in vhba.keys():
                vhba_dict["AssociatedServer_Model"]  = vhba["Profile_AssociatedServer_Model"]
                vhba_dict["AssociatedServer_Name"]   = vhba["Profile_AssociatedServer_Name"]
                vhba_dict["AssociatedServer_Serial"] = vhba["Profile_AssociatedServer_Serial"]
            else:
                vhba_dict["AssociatedServer_Model"]  = ""
                vhba_dict["AssociatedServer_Name"]   = ""
                vhba_dict["AssociatedServer_Serial"] = ""
            for vfc in fi_vfc_data:
                server_serial = vfc["Description"].split(":")[1]
                vhba_name = (vfc["Description"].split()[3]).split(",")[0]
                if "Profile_AssociatedServer" not in vhba.keys():
                    if (vhba["Profile_AssociatedServer_Serial"] == server_serial) and (vhba["Name"] == vhba_name) and (vfc["VfcId"] == vhba["VifId"]):
                        vhba_dict["BoundInterfaceDn"]     = vfc["BoundInterfaceDn"]
                        vhba_dict["Description"]          = vfc["Description"]
                        vhba_dict["Moid"]                 = vfc["Moid"]
                        vhba_dict["FI_AdminEvacState"]    = vfc["NetworkElement_AdminEvacState"]
                        vhba_dict["FI_ManagementMode"]    = vfc["NetworkElement_ManagementMode"]
                        vhba_dict["FI_Model"]             = vfc["NetworkElement_Model"]
                        vhba_dict["FI_OperEvacState"]     = vfc["NetworkElement_OperEvacState"]
                        vhba_dict["FI_Operability"]       = vfc["NetworkElement_Operability"]
                        vhba_dict["FI_Serial"]            = vfc["NetworkElement_Serial"]
                        vhba_dict["FI_SwitchId"]          = vfc["NetworkElement_SwitchId"]
                        vhba_dict["FI_SwitchProfileName"] = vfc["NetworkElement_SwitchProfileName"]
                        vhba_dict["OperReason"]           = vfc["OperReason"]
                        vhba_dict["OperState"]            = vfc["OperState"]
                        vhba_dict["PinnedInterfaceDn"]    = vfc["PinnedInterfaceDn"]
                        vhba_dict["VfcId"]                = vfc["VfcId"]
                else:
                    if (vhba["Name"] == vhba_name) and (vfc["VfcId"] == vhba["VifId"]):
                        vhba_dict["BoundInterfaceDn"]     = vfc["BoundInterfaceDn"]
                        vhba_dict["Description"]          = vfc["Description"]
                        vhba_dict["Moid"]                 = vfc["Moid"]
                        vhba_dict["FI_AdminEvacState"]    = vfc["NetworkElement_AdminEvacState"]
                        vhba_dict["FI_ManagementMode"]    = vfc["NetworkElement_ManagementMode"]
                        vhba_dict["FI_Model"]             = vfc["NetworkElement_Model"]
                        vhba_dict["FI_OperEvacState"]     = vfc["NetworkElement_OperEvacState"]
                        vhba_dict["FI_Operability"]       = vfc["NetworkElement_Operability"]
                        vhba_dict["FI_Serial"]            = vfc["NetworkElement_Serial"]
                        vhba_dict["FI_SwitchId"]          = vfc["NetworkElement_SwitchId"]
                        vhba_dict["FI_SwitchProfileName"] = vfc["NetworkElement_SwitchProfileName"]
                        vhba_dict["OperReason"]           = vfc["OperReason"]
                        vhba_dict["OperState"]            = vfc["OperState"]
                        vhba_dict["PinnedInterfaceDn"]    = vfc["PinnedInterfaceDn"]
                        vhba_dict["VfcId"]                = vfc["VfcId"]
            vhba_data.append(vhba_dict)
        return vhba_data
